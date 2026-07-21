from pathlib import Path

from openpyxl import load_workbook

from catalogos.models import Grado
from miembros.models import Hermano


def importar_hermanos(ruta_excel):

    ruta_excel = Path(ruta_excel)

    if not ruta_excel.exists():
        raise FileNotFoundError(
            f"No existe el archivo: {ruta_excel}"
        )

    wb = load_workbook(
        ruta_excel,
        data_only=True,
    )

    if "MIEMBROS" not in wb.sheetnames:
        raise ValueError(
            "La hoja 'MIEMBROS' no existe."
        )

    ws = wb["MIEMBROS"]

    creados = 0
    actualizados = 0
    omitidos = 0

    for fila in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        numero_control = fila[1]
        nombre = fila[2]
        apellido_paterno = fila[3]
        apellido_materno = fila[4]
        grado_nombre = fila[6]
        fecha_iniciacion = fila[7]
        fecha_exaltacion = fila[8]
        fecha_aumento = fila[9]
        estatus = fila[10]
        telefono = fila[11]
        correo = fila[12]
        fecha_nacimiento = fila[13]
        profesion = fila[14]
        direccion = fila[15]
        observaciones = fila[16]

        if not numero_control:
            continue

        if not nombre:
            omitidos += 1
            continue

        if not apellido_paterno:
            omitidos += 1
            continue

        try:

            grado = Grado.objects.get(
                nombre__iexact=str(grado_nombre).strip()
            )

        except Grado.DoesNotExist:

            print(
                f"[OMITIDO] "
                f"No existe el grado: {grado_nombre}"
            )

            omitidos += 1
            continue

        hermano, creado = Hermano.objects.update_or_create(

            numero_control=str(numero_control),

            defaults={

                "nombre": str(nombre).strip(),

                "apellido_paterno": str(
                    apellido_paterno
                ).strip(),

                "apellido_materno": (
                    str(apellido_materno).strip()
                    if apellido_materno
                    else ""
                ),

                "grado": grado,

                "estatus": (
                    str(estatus).upper().strip()
                    if estatus
                    else "ACTIVO"
                ),

                "telefono": (
                    str(telefono).strip()
                    if telefono
                    else ""
                ),

                "correo": (
                    str(correo).strip()
                    if correo
                    else ""
                ),

                "fecha_nacimiento": fecha_nacimiento,

                "fecha_iniciacion": fecha_iniciacion,

                "fecha_aumento": fecha_aumento,

                "fecha_exaltacion": fecha_exaltacion,

                "profesion": (
                    str(profesion).strip()
                    if profesion
                    else ""
                ),

                "direccion": (
                    str(direccion).strip()
                    if direccion
                    else ""
                ),

                "observaciones": (
                    str(observaciones).strip()
                    if observaciones
                    else ""
                ),

            },

        )

        if creado:
            creados += 1
        else:
            actualizados += 1

    print()
    print("===================================")
    print("IMPORTACIÓN DE HERMANOS FINALIZADA")
    print("===================================")
    print(f"Creados      : {creados}")
    print(f"Actualizados : {actualizados}")
    print(f"Omitidos     : {omitidos}")
    print("===================================")
