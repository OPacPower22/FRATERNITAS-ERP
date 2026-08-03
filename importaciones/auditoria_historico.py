"""
Genera el libro de auditoría del histórico de Tesorería.

Uso:
    python importaciones/auditoria_historico.py \
        importaciones/tesoreria/MOVIMIENTOS_2026.xls \
        AUDITORIA_HISTORICO_2026.xlsx

Produce tres hojas:
    RESUMEN      conciliación mes a mes contra los totales del Excel
    MOVIMIENTOS  las 113 partidas normalizadas, listas para importar
    INCIDENCIAS  filas que requieren decisión antes de importar
"""

import sys
import unicodedata
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(__import__("pathlib").Path(__file__).resolve().parent))

from historico import FONDOS, leer_historico, normalizar, totalizar


FUENTE = "Arial"
GRANATE = "8B0000"
GRIS = "F2F2F2"

BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


# Padrón tomado de la hoja MEMBRESIA del propio histórico.
PADRON = [
    "DAVID GARCIA MELGOZA",
    "GABINO S. SANTAMARIA DELGADO",
    "JORGE L. MENDOZA SORIANO",
    "LEOPOLDO GARCIA MELGOZA",
    "LUIS RODRIGUEZ MATA",
    "L. ENRIQUE RAMON VILABOA",
    "AULIO C. IBARRA ZAVALA",
    "GUILLERMO F. MOTOLINIA SANCHEZ",
    "JOSE A. GONZALEZ SOLIS",
    "J. OMAR PACHECO TAPIA",
    "RAFAEL MARTINEZ DE JESUS",
    "VICTOR H. AMBROSIO GUEVARA",
    "ALEXIS GARCIA ROBLES",
    "ANGEL R. GUTIERREZ ALVAREZ",
    "A. RICARDO GUTIERREZ CONTRERAS",
    "JOSMAR G. CANCHE HERMANDEZ",
    "L. ELISEO JUAREZ MORALES",
    "L. MANUEL SANCHEZ TELLEZ",
]

# Descripciones que no corresponden a un Hermano.
NO_NOMINATIVAS = (
    "SACO",
    "BENEFICENCIA",
    "DEPOSITO",
    "APORTACION",
    "DONATIVO",
    "RIFA",
    "VENTA",
)

# Partidas que son traspasos entre cuentas, no ingreso ni egreso real.
TRASPASOS = ("DEPOSITO DE CAPITAS A BANCO",)


def tokens(nombre):
    """Palabras significativas, ignorando iniciales y partículas."""

    limpio = normalizar(nombre).replace(".", " ")

    return {
        palabra
        for palabra in limpio.split()
        if len(palabra) > 3 and palabra not in {"JOSE", "LUIS"}
    }


def emparejar(descripcion):
    """
    Busca el Hermano del padrón que mejor coincide.

    Retorna (nombre, coincidencias).
    """

    objetivo = tokens(descripcion)

    mejor = ("", 0)

    for candidato in PADRON:
        comunes = len(tokens(candidato) & objetivo)
        if comunes > mejor[1]:
            mejor = (candidato, comunes)

    return mejor


def clasificar(partida):
    """
    Determina la naturaleza de la partida y las incidencias.

    Retorna (hermano, confianza, incidencias)
    """

    descripcion = partida["descripcion_normalizada"]
    incidencias = []

    if any(clave in descripcion for clave in TRASPASOS):
        return "", "TRASPASO", [
            "Traspaso entre cuentas: no debe contarse como ingreso ni egreso."
        ]

    if partida["tipo"] == "E":
        return "", "EGRESO", (
            [] if partida["fecha"] else ["Sin fecha registrada."]
        )

    if any(clave in descripcion for clave in NO_NOMINATIVAS):
        return "", "NO NOMINATIVA", (
            [] if partida["fecha"] else ["Sin fecha registrada."]
        )

    nombre, coincidencias = emparejar(descripcion)

    if coincidencias >= 2:
        confianza = "ALTA"
    elif coincidencias == 1:
        confianza = "MEDIA"
        incidencias.append(
            f"Coincidencia parcial con «{nombre}»: confirmar."
        )
    else:
        confianza = "SIN COINCIDENCIA"
        nombre = ""
        incidencias.append(
            "No corresponde a ningún Hermano del padrón."
        )

    if not partida["fecha"]:
        incidencias.append("Sin fecha registrada.")

    if Decimal(partida["total_declarado"]) not in (
        Decimal("0.00"),
        Decimal(partida["total"]),
    ):
        incidencias.append(
            f"Total del Excel ({partida['total_declarado']}) distinto "
            f"de la suma de fondos ({partida['total']})."
        )

    return nombre, confianza, incidencias


def encabezar(hoja, fila, titulos, anchos):
    for columna, (titulo, ancho) in enumerate(zip(titulos, anchos), start=1):
        celda = hoja.cell(row=fila, column=columna, value=titulo)
        celda.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=GRANATE)
        celda.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        celda.border = BORDE
        hoja.column_dimensions[get_column_letter(columna)].width = ancho
    hoja.freeze_panes = hoja.cell(row=fila + 1, column=1)


def construir(ruta_xls, ruta_salida):

    bloques = leer_historico(ruta_xls)

    libro = Workbook()

    # ------------------------------------------------------------------
    # RESUMEN
    # ------------------------------------------------------------------
    hoja = libro.active
    hoja.title = "RESUMEN"

    hoja["A1"] = "AUDITORÍA DEL HISTÓRICO DE TESORERÍA 2026"
    hoja["A1"].font = Font(name=FUENTE, bold=True, size=14, color=GRANATE)
    hoja["A2"] = f"Archivo de origen: {ruta_xls}"
    hoja["A2"].font = Font(name=FUENTE, italic=True, size=9)

    titulos = [
        "Periodo", "Partidas ingreso", "Partidas egreso",
        "Ingresos ($)", "Egresos ($)", "Neto del mes ($)",
        "Acumulado ($)", "Coincide con el Excel",
    ]
    encabezar(hoja, 4, titulos, [14, 16, 16, 14, 14, 16, 16, 20])

    fila = 5
    primera = fila

    for bloque in bloques:

        ingresos = totalizar(bloque["ingresos"])
        egresos = totalizar(bloque["egresos"])

        declarados_ingreso = bloque["totales_ingresos_declarados"] or {}
        declarados_egreso = bloque["totales_egresos_declarados"] or {}

        cuadra = all(
            declarados_ingreso.get(nombre, Decimal("0.00")) == ingresos[nombre]
            for nombre, _ in FONDOS
        ) and all(
            declarados_egreso.get(nombre, Decimal("0.00")) == egresos[nombre]
            for nombre, _ in FONDOS
        )

        hoja.cell(row=fila, column=1, value=bloque["etiqueta"])
        hoja.cell(row=fila, column=2, value=len(bloque["ingresos"]))
        hoja.cell(row=fila, column=3, value=len(bloque["egresos"]))
        hoja.cell(row=fila, column=4,
                  value=float(sum(ingresos.values(), Decimal("0.00"))))
        hoja.cell(row=fila, column=5,
                  value=float(sum(egresos.values(), Decimal("0.00"))))
        hoja.cell(row=fila, column=6, value=f"=D{fila}-E{fila}")

        if fila == primera:
            hoja.cell(row=fila, column=7, value=f"=F{fila}")
        else:
            hoja.cell(row=fila, column=7, value=f"=G{fila - 1}+F{fila}")

        celda = hoja.cell(row=fila, column=8,
                          value="Sí" if cuadra else "NO — revisar")
        celda.font = Font(name=FUENTE, size=10,
                          bold=not cuadra,
                          color="0F5132" if cuadra else "B02A37")

        fila += 1

    hoja.cell(row=fila, column=1, value="TOTAL").font = Font(
        name=FUENTE, bold=True, size=10
    )
    for columna in (2, 3, 4, 5):
        letra = get_column_letter(columna)
        celda = hoja.cell(
            row=fila,
            column=columna,
            value=f"=SUM({letra}{primera}:{letra}{fila - 1})",
        )
        celda.font = Font(name=FUENTE, bold=True, size=10)
    hoja.cell(row=fila, column=6,
              value=f"=D{fila}-E{fila}").font = Font(name=FUENTE, bold=True,
                                                     size=10)

    for renglon in hoja.iter_rows(min_row=5, max_row=fila, max_col=8):
        for celda in renglon:
            celda.border = BORDE
            if celda.font.name != FUENTE:
                celda.font = Font(name=FUENTE, size=10)
            if celda.column in (4, 5, 6, 7):
                celda.number_format = '$#,##0.00;($#,##0.00);-'

    nota = fila + 2
    hoja.cell(row=nota, column=1,
              value="Los totales calculados se contrastan contra los totales "
                    "escritos en el propio archivo y contra la hoja "
                    "INFORME MENS.").font = Font(name=FUENTE, italic=True,
                                                  size=9)

    # ------------------------------------------------------------------
    # MOVIMIENTOS
    # ------------------------------------------------------------------
    hoja = libro.create_sheet("MOVIMIENTOS")

    titulos = [
        "Fila Excel", "Periodo", "Tipo", "Referencia", "Descripción original",
        "Hermano propuesto", "Confianza", "Fecha",
        "Cápitas", "Aniversario", "Saco Ben.", "Taller AJEF", "Otros",
        "Total", "Nota del Excel", "Incidencias",
    ]
    encabezar(hoja, 1, titulos,
              [10, 10, 8, 12, 32, 30, 16, 12,
               12, 12, 12, 12, 12, 12, 28, 44])

    fila = 2
    incidencias_globales = []

    for bloque in bloques:
        for partida in bloque["ingresos"] + bloque["egresos"]:

            hermano, confianza, incidencias = clasificar(partida)

            hoja.cell(row=fila, column=1, value=partida["fila_excel"])
            hoja.cell(row=fila, column=2, value=partida["periodo"])
            hoja.cell(row=fila, column=3,
                      value="Ingreso" if partida["tipo"] == "I" else "Egreso")
            hoja.cell(row=fila, column=4, value=partida["referencia"])
            hoja.cell(row=fila, column=5, value=partida["descripcion"])
            hoja.cell(row=fila, column=6, value=hermano)

            celda = hoja.cell(row=fila, column=7, value=confianza)
            if confianza in ("MEDIA", "SIN COINCIDENCIA", "TRASPASO"):
                celda.fill = PatternFill("solid", fgColor="FFF3CD")

            hoja.cell(row=fila, column=8, value=partida["fecha"])

            for desplazamiento, (nombre, _) in enumerate(FONDOS):
                hoja.cell(row=fila, column=9 + desplazamiento,
                          value=float(partida[nombre]))

            hoja.cell(row=fila, column=14, value=f"=SUM(I{fila}:M{fila})")
            hoja.cell(row=fila, column=15, value=partida["nota"])
            hoja.cell(row=fila, column=16, value=" ".join(incidencias))

            if incidencias:
                incidencias_globales.append(
                    (
                        partida["fila_excel"],
                        partida["periodo"],
                        partida["descripcion"],
                        confianza,
                        " ".join(incidencias),
                    )
                )

            fila += 1

    ultima = fila - 1
    hoja.cell(row=fila, column=5, value="TOTAL").font = Font(
        name=FUENTE, bold=True, size=10)
    for columna in range(9, 15):
        letra = get_column_letter(columna)
        celda = hoja.cell(row=fila, column=columna,
                          value=f"=SUM({letra}2:{letra}{ultima})")
        celda.font = Font(name=FUENTE, bold=True, size=10)

    for renglon in hoja.iter_rows(min_row=2, max_row=fila, max_col=16):
        for celda in renglon:
            celda.border = BORDE
            if not celda.font.bold:
                celda.font = Font(name=FUENTE, size=10)
            if 9 <= celda.column <= 14:
                celda.number_format = '$#,##0.00;($#,##0.00);-'
            if celda.column == 8:
                celda.number_format = "DD/MM/YYYY"
            if celda.column in (5, 15, 16):
                celda.alignment = Alignment(wrap_text=True, vertical="top")

    hoja.auto_filter.ref = f"A1:P{ultima}"

    # ------------------------------------------------------------------
    # INCIDENCIAS
    # ------------------------------------------------------------------
    hoja = libro.create_sheet("INCIDENCIAS")

    encabezar(hoja, 1,
              ["Fila Excel", "Periodo", "Descripción", "Clasificación",
               "Qué hay que resolver"],
              [12, 12, 34, 20, 70])

    for indice, registro in enumerate(incidencias_globales, start=2):
        for columna, valor in enumerate(registro, start=1):
            celda = hoja.cell(row=indice, column=columna, value=valor)
            celda.font = Font(name=FUENTE, size=10)
            celda.border = BORDE
            if columna in (3, 5):
                celda.alignment = Alignment(wrap_text=True, vertical="top")

    libro.save(ruta_salida)

    return {
        "bloques": len(bloques),
        "partidas": ultima - 1,
        "incidencias": len(incidencias_globales),
    }


if __name__ == "__main__":

    origen = sys.argv[1]
    destino = sys.argv[2] if len(sys.argv) > 2 else "AUDITORIA_HISTORICO.xlsx"

    resumen = construir(origen, destino)

    print(f"Bloques mensuales : {resumen['bloques']}")
    print(f"Partidas          : {resumen['partidas']}")
    print(f"Incidencias       : {resumen['incidencias']}")
    print(f"Archivo generado  : {destino}")
