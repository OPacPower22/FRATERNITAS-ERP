from decimal import Decimal

from openpyxl import load_workbook

from catalogos.models import ConceptoContable
from miembros.models import Hermano
from negocio.models import Obligacion

CUOTA_MENSUAL = Decimal("280.00")


def importar_obligaciones(ruta_excel):

    concepto = ConceptoContable.objects.get(
        nombre="Cuota Ordinaria"
    )

    wb = load_workbook(
        ruta_excel,
        data_only=True,
    )

    ws = wb["MEMBRESIA"]

    creadas = 0

    for fila in ws.iter_rows(
        min_row=4,
        values_only=True,
    ):

        numero_control = fila[0]
        meses_adeudo = fila[4]

        if not numero_control:
            continue

        try:

            hermano = Hermano.objects.get(
                numero_control=str(numero_control)
            )

        except Hermano.DoesNotExist:

            continue

        try:

            meses = int(meses_adeudo)

        except (TypeError, ValueError):

            meses = 0

        for i in range(meses):

            Obligacion.objects.get_or_create(

                hermano=hermano,

                concepto=concepto,

                periodo=f"Adeudo {i+1}",

                defaults={

                    "importe": CUOTA_MENSUAL,

                    "saldo_pendiente": CUOTA_MENSUAL,

                    "estado": "PENDIENTE",

                    "fecha_vencimiento": "2026-01-01",

                },

            )

            creadas += 1

    print(f"Obligaciones creadas: {creadas}")