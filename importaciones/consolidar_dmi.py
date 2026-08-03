"""
Consolida el Documento Maestro Institucional (DMI).

Toma como base el DMI del repositorio (que ya tiene la estructura
que el código espera) y le aplica las correcciones que se derivan
de dos fuentes:

  - TESORERIA WEB APP ... .xlsx  (tarifas originales de la Logia)
  - MOVIMIENTOS_2026.xls         (operación real del ejercicio)

Cada celda modificada queda marcada en amarillo y documentada en
la hoja CORRECCIONES.

Uso:
    python importaciones/consolidar_dmi.py BASE.xlsx SALIDA.xlsx
"""

import sys

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill


FUENTE = "Arial"
AMARILLO = PatternFill("solid", fgColor="FFFF00")
GRANATE = "8B0000"


# ----------------------------------------------------------------------
# Correcciones a aplicar
# ----------------------------------------------------------------------

# hoja -> {clave en la primera columna: (columna, valor nuevo, motivo)}

TARIFAS = {
    "MEMBRESIA": (
        3, 120,
        "PAGO A GLUM cobra $120 por Hermano cada mes, no $150. "
        "El libro TESORERIA WEB APP también registra 120.",
    ),
    "APORTACION_FRATERNIDAD": (
        3, 50,
        "La columna ANIVERSARIO del histórico registra $50 en todos "
        "los recibos completos, no $70.",
    ),
}

DISTRIBUCION = {
    "MEMBRESIA": (
        3, 120,
        "Debe coincidir con la tarifa: $120 por Hermano.",
    ),
    "APORTACION_FRATERNIDAD": (
        3, 50,
        "Debe coincidir con la tarifa: $50 por Hermano.",
    ),
}

PARAMETROS = {
    "CAPITA_MENSUAL": (
        2, 350,
        "260 (GLUM) + 90 (local) = 350. Es el importe que el "
        "histórico registra en todos los recibos completos: "
        "285 cápitas + 50 aniversario + 15 taller AJEF.",
    ),
}

# Hermanos que causaron baja durante el ejercicio y no están en el DMI.
BAJAS = [
    {
        "numero_control": 16,
        "nombre": "Josmar Gabriel",
        "apellido_paterno": "Canché",
        "apellido_materno": "Hernández",
        "grado": "Maestro",
        "estado": "BAJA",
        "observaciones": "Baja en abril de 2026 (hoja MEMBRESIA).",
    },
    {
        "numero_control": 17,
        "nombre": "Luis Eliseo",
        "apellido_paterno": "Juárez",
        "apellido_materno": "Morales",
        "grado": "Maestro",
        "estado": "BAJA",
        "observaciones": "Baja en mayo de 2026 (hoja MEMBRESIA).",
    },
    {
        "numero_control": 18,
        "nombre": "Luis Manuel",
        "apellido_paterno": "Sánchez",
        "apellido_materno": "Téllez",
        "grado": "Maestro",
        "estado": "BAJA",
        "observaciones": "Baja en marzo de 2026 (hoja MEMBRESIA).",
    },
]


def _localizar(hoja, clave, columna_clave=2):
    """Devuelve el número de fila cuya columna clave coincide."""

    for fila in range(2, hoja.max_row + 1):
        if str(hoja.cell(row=fila, column=columna_clave).value).strip() == clave:
            return fila

    return None


def _aplicar(hoja, correcciones, columna_clave, bitacora, nombre_hoja):

    for clave, (columna, valor_nuevo, motivo) in correcciones.items():

        fila = _localizar(hoja, clave, columna_clave)

        if fila is None:
            bitacora.append(
                (nombre_hoja, "—", clave, "", "", f"NO ENCONTRADO: {motivo}")
            )
            continue

        celda = hoja.cell(row=fila, column=columna)
        anterior = celda.value

        celda.value = valor_nuevo
        celda.fill = AMARILLO
        celda.comment = Comment(motivo, "Auditoría FRATERNITAS-ERP")

        bitacora.append(
            (nombre_hoja, celda.coordinate, clave, anterior, valor_nuevo, motivo)
        )


def consolidar(ruta_base, ruta_salida):

    libro = load_workbook(ruta_base)
    bitacora = []

    _aplicar(libro["TARIFAS_OBLIGACIONES"], TARIFAS, 2, bitacora,
             "TARIFAS_OBLIGACIONES")

    _aplicar(libro["DISTRIBUCION_CAPITA"], DISTRIBUCION, 2, bitacora,
             "DISTRIBUCION_CAPITA")

    _aplicar(libro["PARAMETROS"], PARAMETROS, 1, bitacora, "PARAMETROS")

    # ------------------------------------------------------------------
    # MIEMBROS: marcar estado de los vigentes y agregar las tres bajas
    # ------------------------------------------------------------------
    hoja = libro["MIEMBROS"]

    encabezados = {
        str(celda.value).strip(): celda.column
        for celda in hoja[1]
        if celda.value
    }

    columna_estado = encabezados.get("ESTADO")
    columna_control = encabezados.get("NUMERO_CONTROL")

    ultima = hoja.max_row

    for fila in range(2, ultima + 1):
        if hoja.cell(row=fila, column=columna_control).value is None:
            continue
        celda = hoja.cell(row=fila, column=columna_estado)
        if not celda.value:
            celda.value = "ACTIVO"
            celda.fill = AMARILLO

    bitacora.append(
        ("MIEMBROS", f"K2:K{ultima}", "ESTADO", "(vacío)", "ACTIVO",
         "El sincronizador no puede distinguir activos de bajas si la "
         "columna viene vacía.")
    )

    fila = ultima + 1

    for baja in BAJAS:
        hoja.cell(row=fila, column=encabezados["NUMERO_CONTROL"],
                  value=baja["numero_control"])
        hoja.cell(row=fila, column=encabezados["NOMBRE"],
                  value=baja["nombre"])
        hoja.cell(row=fila, column=encabezados["APELLIDO_PATERNO"],
                  value=baja["apellido_paterno"])
        hoja.cell(row=fila, column=encabezados["APELLIDO_MATERNO"],
                  value=baja["apellido_materno"])
        hoja.cell(row=fila, column=encabezados["NOMBRE_COMPLETO"],
                  value=(
                      f"{baja['nombre']} {baja['apellido_paterno']} "
                      f"{baja['apellido_materno']}"
                  ))
        hoja.cell(row=fila, column=encabezados["GRADO"], value=baja["grado"])
        hoja.cell(row=fila, column=columna_estado, value=baja["estado"])

        if "OBSERVACIONES" in encabezados:
            hoja.cell(row=fila, column=encabezados["OBSERVACIONES"],
                      value=baja["observaciones"])

        for columna in range(1, max(encabezados.values()) + 1):
            hoja.cell(row=fila, column=columna).fill = AMARILLO

        bitacora.append(
            ("MIEMBROS", f"fila {fila}", baja["nombre"], "(no existía)",
             baja["estado"], baja["observaciones"])
        )

        fila += 1

    # ------------------------------------------------------------------
    # Hoja CORRECCIONES
    # ------------------------------------------------------------------
    if "CORRECCIONES" in libro.sheetnames:
        del libro["CORRECCIONES"]

    hoja = libro.create_sheet("CORRECCIONES", 0)

    hoja["A1"] = "CORRECCIONES APLICADAS AL DMI"
    hoja["A1"].font = Font(name=FUENTE, bold=True, size=14, color=GRANATE)
    hoja["A2"] = (
        "Toda celda modificada está resaltada en amarillo dentro de su "
        "hoja y lleva un comentario con el motivo. "
        "Revise y confirme antes de sincronizar."
    )
    hoja["A2"].font = Font(name=FUENTE, italic=True, size=9)

    titulos = ["Hoja", "Celda", "Clave", "Valor anterior",
               "Valor nuevo", "Motivo"]
    anchos = [24, 12, 26, 16, 14, 80]

    for columna, (titulo, ancho) in enumerate(zip(titulos, anchos), start=1):
        celda = hoja.cell(row=4, column=columna, value=titulo)
        celda.font = Font(name=FUENTE, bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=GRANATE)
        celda.alignment = Alignment(horizontal="center", wrap_text=True)
        hoja.column_dimensions[
            celda.column_letter
        ].width = ancho

    for indice, registro in enumerate(bitacora, start=5):
        for columna, valor in enumerate(registro, start=1):
            celda = hoja.cell(row=indice, column=columna, value=valor)
            celda.font = Font(name=FUENTE, size=10)
            celda.alignment = Alignment(wrap_text=True, vertical="top")

    hoja.freeze_panes = "A5"

    libro.save(ruta_salida)

    return bitacora


if __name__ == "__main__":

    base = sys.argv[1]
    salida = sys.argv[2] if len(sys.argv) > 2 else "DMI_PARAMETROS_2026.xlsx"

    registros = consolidar(base, salida)

    for registro in registros:
        print(f"{registro[0]:22} {registro[1]:>10}  "
              f"{registro[2]:26} {registro[3]} -> {registro[4]}")

    print(f"\nCorrecciones aplicadas: {len(registros)}")
    print(f"Archivo generado      : {salida}")
