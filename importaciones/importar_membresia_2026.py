from pathlib import Path

from openpyxl import load_workbook

from catalogos.models import Grado
from miembros.models import Hermano

GRADOS = (
    ("Aprendiz", "APR", 1),
    ("Compañero", "COMP", 2),
    ("Maestro Mason", "MM", 3),
)

GRADO_POR_DEFECTO = "Maestro Mason"

CONECTORES_APELLIDO = {"DE", "DEL", "LA", "LOS", "LAS", "Y"}


def _asegurar_grados():

    for nombre, abreviatura, orden in GRADOS:

        Grado.objects.update_or_create(
            nombre=nombre,
            defaults={
                "abreviatura": abreviatura,
                "orden": orden,
                "activo": True,
            },
        )


def _separar_nombre(nombre_completo):
    """
    Separa un nombre completo en nombre(s) y apellidos, respetando
    apellidos compuestos con conectores (ej. "MARTINEZ DE JESUS").
    """

    palabras = nombre_completo.split()

    apellidos = []

    while len(apellidos) < 2 and palabras:

        grupo = [palabras.pop()]

        while palabras and palabras[-1].upper().rstrip(".") in CONECTORES_APELLIDO:
            grupo.insert(0, palabras.pop())

        apellidos.insert(0, " ".join(grupo))

    apellido_paterno = apellidos[0] if apellidos else ""
    apellido_materno = apellidos[1] if len(apellidos) > 1 else ""
    nombre = " ".join(palabras)

    return nombre, apellido_paterno, apellido_materno


def importar_membresia(ruta_excel):

    ruta_excel = Path(ruta_excel)

    if not ruta_excel.exists():
        raise FileNotFoundError(
            f"No existe el archivo: {ruta_excel}"
        )

    _asegurar_grados()

    grado = Grado.objects.get(nombre=GRADO_POR_DEFECTO)

    wb = load_workbook(
        ruta_excel,
        data_only=True,
    )

    if "MEMBRESIA" not in wb.sheetnames:
        raise ValueError(
            "La hoja 'MEMBRESIA' no existe."
        )

    ws = wb["MEMBRESIA"]

    creados = 0
    actualizados = 0
    omitidos = 0

    numero_control = 0

    for fila in ws.iter_rows(
        min_row=4,
        values_only=True,
    ):

        nombre_completo = fila[1]

        if not nombre_completo or not str(nombre_completo).strip():
            continue

        numero_control += 1

        nombre_completo = " ".join(str(nombre_completo).split())

        nombre, apellido_paterno, apellido_materno = _separar_nombre(
            nombre_completo
        )

        if not nombre or not apellido_paterno:
            omitidos += 1
            continue

        hermano, creado = Hermano.objects.update_or_create(

            numero_control=str(numero_control),

            defaults={

                "nombre": nombre,

                "apellido_paterno": apellido_paterno,

                "apellido_materno": apellido_materno,

                "grado": grado,

                "estatus": "ACTIVO",

            },

        )

        if creado:
            creados += 1
        else:
            actualizados += 1

    print()
    print("===================================")
    print("IMPORTACIÓN DE MEMBRESÍA 2026 FINALIZADA")
    print("===================================")
    print(f"Creados      : {creados}")
    print(f"Actualizados : {actualizados}")
    print(f"Omitidos     : {omitidos}")
    print("===================================")
