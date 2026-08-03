from collections import defaultdict

from openpyxl import load_workbook

from negocio.services.mid.config import DMI, HOJAS

from negocio.services.mid.sincronizar_grados import (
        sincronizar_grados,
        )

from negocio.services.mid.sincronizar_cargos import (
    sincronizar_cargos,
)

from negocio.services.mid.sincronizar_conceptos import (
    sincronizar_conceptos,
)

from negocio.services.mid.sincronizar_parametros import (
    sincronizar_parametros,
)

def sincronizar_catalogos():

    if not DMI.exists():
        raise FileNotFoundError(
            f"No se encontró el DMI:\n{DMI}"
        )

    libro = load_workbook(
        DMI,
        data_only=True,
    )

    nombre_hoja = HOJAS["catalogos"]

    if nombre_hoja not in libro.sheetnames:
        raise ValueError(
            f"La hoja '{nombre_hoja}' no existe."
        )

    hoja = libro[nombre_hoja]

    catalogos = defaultdict(list)

    # Los encabezados se localizan por nombre y no por posición:
    # el DMI tiene la columna ORDEN entre DESCRIPCION y ACTIVO,
    # mientras que la exportación de la web app no la tiene.
    encabezados = {
        str(celda).strip().upper(): indice
        for indice, celda in enumerate(
            next(hoja.iter_rows(max_row=1, values_only=True))
        )
        if celda
    }

    indice_tipo = encabezados.get("CATALOGO", 0)
    indice_clave = encabezados.get("CLAVE", 1)
    indice_descripcion = encabezados.get("DESCRIPCION", 2)
    indice_activo = encabezados.get("ACTIVO", 3)

    for fila in hoja.iter_rows(
        min_row=2,
        values_only=True,
    ):

        tipo = fila[indice_tipo]
        clave = fila[indice_clave]
        descripcion = fila[indice_descripcion]
        activo = fila[indice_activo]

        if not tipo:
            continue

        catalogos[tipo].append(
            {
                "clave": clave,
                "descripcion": descripcion,
                "activo": activo,
            }
        )

    print()
    print("=" * 60)
    print("CATÁLOGOS ENCONTRADOS")
    print("=" * 60)

    for nombre, elementos in sorted(catalogos.items()):
        print(f"{nombre:<35}{len(elementos):>5}")

    print("=" * 60)

    if "GRADOS_MASONICOS" in catalogos:
          sincronizar_grados(
                 catalogos["GRADOS_MASONICOS"]
           )
    
    if "CARGOS_LOGIA" in catalogos:

           sincronizar_cargos(
                catalogos["CARGOS_LOGIA"]
           )
    
    if "OBLIGACION" in catalogos:

            sincronizar_conceptos(
                catalogos["OBLIGACION"]
            )
    
#   if "PARAMETROS" in catalogos:
#
#         sincronizar_parametros(
#            catalogos["PARAMETROS"]
#         )

    return catalogos
