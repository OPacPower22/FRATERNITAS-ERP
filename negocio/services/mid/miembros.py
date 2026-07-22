from openpyxl import load_workbook

from negocio.services.mid.config import DMI
from miembros.models import Hermano
from catalogos.models import Grado

def leer_miembros():
    """
    Lee la hoja MIEMBROS del DMI.
    """

    wb = load_workbook(
        DMI,
        read_only=True,
        data_only=True,
    )

    ws = wb["MIEMBROS"]

    encabezados = [
        c.value
        for c in next(
            ws.iter_rows(max_row=1)
        )
    ]

    miembros = []

    for fila in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        if all(v is None for v in fila):
            continue

        registro = dict(
            zip(
                encabezados,
                fila,
            )
        )

        miembros.append(registro)

    wb.close()

    return miembros

from miembros.models import Hermano
from catalogos.models import Grado


def sincronizar_miembros():

    datos = leer_miembros()

    creados = 0
    actualizados = 0

    for registro in datos:

        grado = Grado.objects.filter(
            nombre__iexact=registro["GRADO"]
        ).first()

        _, creado = Hermano.objects.update_or_create(
            numero_control=str(registro["NUMERO_CONTROL"]),
            defaults={
                "nombre": registro["NOMBRE"] or "",
                "apellido_paterno": registro["APELLIDO_PATERNO"] or "",
                "apellido_materno": registro["APELLIDO_MATERNO"] or "",
                "grado": grado,
                "fecha_iniciacion": registro["FECHA_INICIACION"],
                "fecha_exaltacion": registro["FECHA_EXALTACION"],
                "telefono": registro["TELEFONO"] or "",
                "correo": registro["CORREO"] or "",
                "profesion": registro["PROFESION"] or "",
                "direccion": registro["DIRECCION"] or "",
                "observaciones": registro["OBSERVACIONES"] or "",
            },
        )

        if creado:
            creados += 1
        else:
            actualizados += 1

    print("\nMIEMBROS")
    print(f"Creados.....: {creados}")
    print(f"Actualizados: {actualizados}")
