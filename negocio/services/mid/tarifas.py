from openpyxl import load_workbook

from negocio.services.mid.config import DMI, HOJAS


def leer_tarifas():

    if not DMI.exists():
        raise FileNotFoundError(
            f"No se encontró el DMI:\n{DMI}"
        )

    libro = load_workbook(
        DMI,
        data_only=True,
    )

    nombre_hoja = HOJAS["tarifas"]

    if nombre_hoja not in libro.sheetnames:
        raise ValueError(
            f"La hoja '{nombre_hoja}' no existe."
        )

    hoja = libro[nombre_hoja]

    tarifas = []

    for fila in hoja.iter_rows(
        min_row=2,
        values_only=True,
    ):

        if not fila[0]:
            continue

        tarifas.append(
            {
                "id_tarifa": fila[0],
                "concepto": fila[1],
                "importe": fila[2],
                "obligatorio": fila[3],
                "estado": fila[4],
            }
        )

    return tarifas
