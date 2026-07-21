from openpyxl import load_workbook

from negocio.services.mid.config import DMI, HOJAS
from negocio.services.mid.sincronizar_parametros import (
    sincronizar_parametros,
)


def sincronizar_parametros_dmi():

    if not DMI.exists():
        raise FileNotFoundError(
            f"No se encontró el DMI:\n{DMI}"
        )

    libro = load_workbook(
        DMI,
        data_only=True,
    )

    nombre_hoja = HOJAS["parametros"]

    if nombre_hoja not in libro.sheetnames:
        raise ValueError(
            f"La hoja '{nombre_hoja}' no existe."
        )

    hoja = libro[nombre_hoja]

    parametros = []

    for fila in hoja.iter_rows(
        min_row=2,
        values_only=True,
    ):

        parametro = fila[0]
        valor = fila[1]
        descripcion = fila[2]
        modificable = fila[3]

        if not parametro:
            continue

        parametros.append(
            {
                "parametro": parametro,
                "valor": valor,
                "descripcion": descripcion,
                "modificable": modificable,
            }
        )

    print()

    print("=" * 60)
    print("PARAMETROS")
    print("=" * 60)
    print(f"Registros encontrados: {len(parametros)}")
    print("=" * 60)

    sincronizar_parametros(parametros)
