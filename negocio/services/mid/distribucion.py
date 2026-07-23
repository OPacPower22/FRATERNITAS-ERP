from openpyxl import load_workbook

from negocio.services.mid.config import (
    DMI,
    HOJAS,
)

from negocio.services.mid.sincronizar_distribucion import (
    sincronizar_distribucion,
)

def leer_distribucion():

    libro = load_workbook(
        DMI,
        data_only=True,
    )

    hoja = libro[
        HOJAS["distribucion"]
    ]

    registros = []

    encabezados = [
        str(celda.value).strip().lower()
        for celda in hoja[1]
    ]

    for fila in hoja.iter_rows(
        min_row=2,
        values_only=True,
    ):

        if all(
            valor is None
            for valor in fila
        ):
            continue

        registros.append(
            dict(
                zip(
                    encabezados,
                    fila,
                )
            )
        )

    return registros

def sincronizar_distribucion_dmi():

    distribucion = leer_distribucion()

    sincronizar_distribucion(
        distribucion
    )