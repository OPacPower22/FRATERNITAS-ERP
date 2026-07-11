from pathlib import Path

from openpyxl import load_workbook


class ImportadorExcel:

    def obtener_informacion_libro(self):

        ruta = (
            Path(__file__)
            .resolve()
            .parent.parent.parent
            / "importaciones"
            / "tesoreria"
            / "TESORERIA_MOVIMIENTOS_2026.xlsx"
        )

        if not ruta.exists():
            return {
                "ok": False,
                "mensaje": "No se encontró el archivo de importación.",
            }

        libro = load_workbook(
            filename=ruta,
            read_only=True,
            data_only=True,
        )

        return {
            "ok": True,
            "archivo": str(ruta),
            "hojas": libro.sheetnames,
        }

    def obtener_hoja_movimientos(self):

        informacion = self.obtener_informacion_libro()

        if not informacion["ok"]:
            return informacion

        ruta = Path(informacion["archivo"])

        libro = load_workbook(
            filename=ruta,
            read_only=True,
            data_only=True,
        )

        hoja = libro["INGRESOS Y EGRESOS"]

        return {
            "ok": True,
            "hoja": hoja,
        }

    def obtener_filas(self):

        resultado = self.obtener_hoja_movimientos()

        if not resultado["ok"]:
            return resultado

        hoja = resultado["hoja"]

        return {
            "ok": True,
            "filas": hoja.max_row,
            "columnas": hoja.max_column,
        }

    def obtener_primeras_filas(self, cantidad=15):

        resultado = self.obtener_hoja_movimientos()

        if not resultado["ok"]:
            return resultado

        hoja = resultado["hoja"]

        filas = []

        for fila in hoja.iter_rows(
            min_row=1,
            max_row=cantidad,
            values_only=True,
        ):
            filas.append(fila)

        return {
            "ok": True,
            "filas": filas,
        }

    def obtener_encabezados(self):

        resultado = self.obtener_hoja_movimientos()

        if not resultado["ok"]:
            return resultado

        hoja = resultado["hoja"]

        encabezados = []

        for celda in hoja[7]:
            encabezados.append(celda.value)

        return {
            "ok": True,
            "encabezados": encabezados,
        }

    def obtener_primer_movimiento(self):

        resultado = self.obtener_hoja_movimientos()

        if not resultado["ok"]:
            return resultado

        hoja = resultado["hoja"]

        fila = next(
            hoja.iter_rows(
                min_row=8,
                max_row=8,
                values_only=True,
            )
        )

        return {
            "ok": True,
            "movimiento": {
                "recibo": fila[0],
                "miembro": fila[1],
                "capitas": fila[4],
                "aniversario": fila[5],
                "saco_beneficencia": fila[6],
                "taller_bj": fila[7],
                "otros": fila[8],
                "total": fila[9],
                "fecha": fila[10],
                "observaciones": fila[11],
            },
        }

    def obtener_movimientos(self):

        resultado = self.obtener_hoja_movimientos()

        if not resultado["ok"]:
            return resultado
        hoja = resultado["hoja"]

        movimientos = []

        for fila in hoja.iter_rows(
            min_row=8,
            values_only=True,
        ):

            if not any(fila):
                continue

            miembro = fila[1]

            if miembro is None:
                continue

            miembro = str(miembro).strip().upper()

            if miembro == "M I E M B R O":
                continue

            if miembro == "C O N C E P T O":
                continue

            movimientos.append(fila)

        return {
            "ok": True,
            "movimientos": movimientos,
        }